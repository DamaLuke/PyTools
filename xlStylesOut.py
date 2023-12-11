import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
import pandas as pd
from tkinter import messagebox as msg
from rich.console import Console


# 默认格式
title_font = Font(
    name="MYingHei_18030_C-Medium", size=10, bold=True, color="00FFFFFF"
)
title_fill = PatternFill(
    fill_type="solid", start_color="3D5AAE", end_color="3D5AAE"
)
body_font = Font(
    name="MYingHei_18030_C-Medium", size=9, bold=False, color="000000"
)
body_fill = PatternFill(
    fill_type="solid", start_color="00FFFFFF", end_color="00FFFFFF"
)
thin = Side(border_style="thin", color="000000")
border = Border(left = thin, right = thin, top = thin, bottom = thin)
align = Alignment(horizontal="center", vertical="center", wrapText=True)    

title_namestyle = NamedStyle(name = "default title format",
                               font = title_font,
                               fill = title_fill,
                               border = border,
                               alignment = align)
body_namestyle = NamedStyle(name = "default body format",
                               font = body_font,
                               fill = body_fill,
                               border = border,
                               alignment = align)
    
    
def write_excel(path: str, 
                dfs: dict[str, object],
                headers = 1, 
                titleStyle = title_namestyle,
                bodyStyle = body_namestyle,                
                isCompeleted = False) -> None:
    """Write muti-dataframes into excel file and formatted that. 
    
    Args:
        path: path to write the spreedsheets(str).
        dfs: A dictionary of (sheet names:keys, DataFrames:values) in pairs.
        headers: The number of rows of title.
        titleStyle: The style of title. It's a openpyxl.styles.NamedStyle class.
        bodyStyle: The style of body. It's a openpyxl.styles.NamedStyle class.
        isCompeleted: Whether to showinfo of is completed."""
        
    def completedInfo():
        msg.showinfo("Completed", "已完成！")    
    
    console = Console()
    with console.status("Writing Excel to {}".format(path)):  
        with pd.ExcelWriter(path) as writer:
            for sheet, df in dfs.items():
                df.to_excel(writer, sheet_name=sheet, index=False)
                df = df.drop_duplicates() 
        wb = openpyxl.load_workbook(path)
        wb.add_named_style(bodyStyle)
        for ws in wb:
            for col in range(ws.min_column, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True
                for i in range(1, headers + 1):
                    cell = ws.cell(i, col)
                    cell.style = titleStyle
            ws.auto_filter.ref = f"A{headers}:{get_column_letter(col)}{headers}"
        wb.save(path)
    
    print("文件路径为 " + path)
    if isCompeleted == True:
        completedInfo()       
