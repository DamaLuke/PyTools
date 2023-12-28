import pandas as pd
import openpyxl
import os
import copy
import itertools
        
"""
使用excelCompare输入参数创建实例，再调用compare_sheet()即可进行比对

Parameters
    1.newfile_path: 本次excel文件的绝对路径
    2.oldfile_path: 上一次excel文件的绝对路径
    3.compfile_path: 输出比对文件的绝对路径
    4.key_names: 比对关键变量，需要用"|"隔开。如 序号|访视名称|日期类型|受试者代码|表名称|检查项目
    5.pre_names: 批注保留变量，需要用"|"隔开。如 审核人员|备注

Example 
    rootpath=os.path.dirname(__file__)                                     
    newfile_path = os.path.join(rootpath,'MA-ASCT-II-002_OffLineListing_2023-12-28.xlsx')
    oldfile_path = os.path.join(rootpath,'MA-ASCT-II-002_OffLineListing_2023-12-21.xlsx')
    compfile_path = os.path.join(rootpath,'CompareResults.xlsx')


    key_names = "序号|访视名称|日期类型|受试者代码|表名称|检查项目" 
    pre_names = "审核人员|备注"                         
        
    xlcomp = excelCompare(newfile_path=newfile_path,
                        oldfile_path=oldfile_path,
                        compfile_path=compfile_path,
                        key_names=key_names,
                        pre_names=pre_names).compare_sheet()
"""  
def create_Row_status(ws):
    first_col = ws['A']
    copied_style = [cell.style for cell in first_col]
    ws.insert_cols(0) 
    for i, style in enumerate(copied_style, start=1):
        ws.cell(i, 1).style = style
    ws.cell(1,1).value = 'Row Status'   

class rowdata():#以row为单位获取主键值、注释值、比对值等信息
    def __init__(self,ws,row,key_indexes,pre_indexs):
        self.ws = ws
        self.row = row
        self.key_indexes = key_indexes
        self.pre_indexes = pre_indexs
        self.key_values = [ self.ws.cell(self.row, col).value for col in range(1,self.ws.max_column + 1) if col in self.key_indexes ]
        self.pre_values = { self.ws.cell(1,col).value : self.ws.cell(self.row, col).value for col in range(1,self.ws.max_column + 1) if col in self.pre_indexes }
        self.comp_values = { self.ws.cell(1,col).value : self.ws.cell(self.row, col).value for col in range(1,self.ws.max_column + 1) if col not in set(self.key_indexes + self.pre_indexes) }
        self.keycell_cord = { self.ws.cell(1,col).value : self.ws.cell(self.row, col).coordinate for col in range(1,self.ws.max_column + 1) if col in self.key_indexes }
        self.precell_cord = { self.ws.cell(1,col).value : self.ws.cell(self.row, col).coordinate for col in range(1,self.ws.max_column + 1) if col in self.pre_indexes }
        self.compcell_cord = { self.ws.cell(1,col).value : self.ws.cell(self.row, col).coordinate for col in range(1,self.ws.max_column + 1) if col not in set(self.key_indexes + self.pre_indexes) } 
        
#根据title返回相应列位置
def get_column_index(file_path, sheet_name, column_names):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_names = column_names.split("|")
    column_indexes = []
    for column_name in column_names:
        if  column_name in df.columns:
            column_indexes.append( df.columns.get_loc(column_name) + 1 )
    return column_indexes

#获取excel文件中所有的sheet名称
def get_sheet_names(file_path):
    xl = pd.ExcelFile(file_path)
    return xl.sheet_names    

#删除Row Status列
def delete_status_col(wb):
    for sheet in wb:
        for col in sheet.iter_cols():
            for cell in col:
                if sheet.cell(1,cell.column).value == 'Row Status':
                    sheet.delete_cols(cell.column)        
    return wb
    
#返回共同存在的sheet，仅存于old的sheet，仅存于new的sheet     
def match_sheets(old,new):
    set1 = set(old)
    set2 = set(new)
    common = set1 & set2
    only_in_old = set1 - set2
    only_in_new = set2 - set1
    return common, only_in_old, only_in_new        
                    
                     
class excelCompare():          
    def __init__(self,newfile_path,oldfile_path,compfile_path,key_names,pre_names,startrow = 2):
        self.newfile_path = newfile_path
        self.oldfile_path = oldfile_path
        self.compfile_path = compfile_path
        self.key_names = key_names
        self.pre_names = pre_names    
        self.oldsheet_name = get_sheet_names(self.oldfile_path)
        self.newsheet_name = get_sheet_names(self.newfile_path)    
        self.startrow = startrow     

    def create_res_wb(self): #复制newfile副本，随后在副本上执行相关操作
        new_wb = openpyxl.load_workbook(self.newfile_path)
        res_wb = new_wb      
        res_wb.save(self.compfile_path)   
        return res_wb          

    #commonsheet进行比对，分别输出changed/unchanged/newly added/deleted四种状态 
    def compare_sheet(self):
        matchsheet = match_sheets(self.oldsheet_name,self.newsheet_name)[0]   
        only_in_oldsheet = match_sheets(self.oldsheet_name,self.newsheet_name)[1]   
        only_in_newsheet = match_sheets(self.oldsheet_name,self.newsheet_name)[2]   
        old_wb = openpyxl.load_workbook(self.oldfile_path,data_only=True)
        old_wb = delete_status_col(old_wb)
        new_wb = self.create_res_wb()
        new_wb = delete_status_col(new_wb)
        startrow = self.startrow
        
        for item in matchsheet:
            sheet_name=item  
            
            oldkey_indexes = get_column_index(self.oldfile_path, sheet_name, self.key_names)
            oldpre_indexes = get_column_index(self.oldfile_path, sheet_name, self.pre_names)
            newkey_indexes = get_column_index(self.newfile_path, sheet_name, self.key_names)
            newpre_indexes = get_column_index(self.newfile_path, sheet_name, self.pre_names)

            old_ws = old_wb[sheet_name] 
            new_ws = new_wb[sheet_name]  
            status_rowflag = {i : '' for i in range(startrow, new_ws.max_row + 1)}
            oldkeydict = {}
            newkeydict = {}
            
            print('comparing matched sheet: %s' % (sheet_name))
            
            for row in range(startrow, old_ws.max_row + 1 ):               
                old = rowdata(old_ws,row,oldkey_indexes,oldpre_indexes) 
                oldkeydict[tuple(old.key_values)] = row 
                
            for row in range(startrow, new_ws.max_row + 1 ): 
                new = rowdata(new_ws,row,newkey_indexes,newpre_indexes)
                newkeydict[tuple(new.key_values)] = row
                
            for oldkeyvalue,newkeyvalue in itertools.product(oldkeydict,newkeydict):  
                if oldkeyvalue == newkeyvalue:
                    oldrow = oldkeydict[oldkeyvalue]
                    newrow = newkeydict[newkeyvalue]
                    oldrowdata = rowdata(old_ws,oldrow,oldkey_indexes,oldpre_indexes)
                    newrowdata = rowdata(new_ws,newrow,newkey_indexes,newpre_indexes)
                    
                    for cell in new_ws[newrow]: #保留上一次的comment信息
                        prenames = new_ws.cell(1, cell.column).value
                        oldprevalue = oldrowdata.pre_values.get(prenames)
                        newprevalue = oldrowdata.pre_values.get(prenames)
                        if oldprevalue != newprevalue:
                            new_ws.cell(newrow,cell.column).value = old_ws.cell(row,cell.column).value
                                
                    for key in newrowdata.comp_values.keys():#判断changed的数据，用批注显示数据的变化 
                        if oldrowdata.comp_values.get(key) != newrowdata.comp_values.get(key):
                            print('key value:' + str(oldkeyvalue))
                            status_rowflag[newrowdata.row] = 'changed'
                            cord = newrowdata.compcell_cord.get(key)
                            cell = new_ws[cord]
                            yellow_fill=openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00",fill_type='solid') 
                            cell.fill = yellow_fill
                            comment = openpyxl.comments.Comment('Previous Value: \n' + str(oldrowdata.comp_values.get(key)) , 'auto_user')
                            cell.comment = comment    
                            print('value: %s' % (str(newrowdata.comp_values.get(key))))
                            print('previous value: %s' % (str(oldrowdata.comp_values.get(key))))  
                                                                                                                    
               
            # 复制第一列的内容，清除内容并保留格式，作为Row Status
            create_Row_status(new_ws)    
            
            for key in oldkeydict.keys() - newkeydict.keys():
                new_ws.insert_rows(new_ws.max_row + 1)
                maxrow = new_ws.max_row + 1

                for oldcell in old_ws[oldkeydict.get(key)]:
                    new_ws.cell(maxrow,oldcell.column + 1).value = oldcell.value
                    new_ws.cell(maxrow,oldcell.column + 1).font = copy.copy(oldcell.font)                
                    new_ws.cell(maxrow,oldcell.column + 1).border = copy.copy(oldcell.border)  
                    
                status_rowflag[maxrow]='deleted'
                                        
            for key in newkeydict.keys() - oldkeydict.keys():                    
                status_rowflag[newkeydict.get(key)]='newly added'    
                
            for key in status_rowflag:
                new_ws.cell(key,1).value = status_rowflag.get(key)    
            
            for cell in new_ws['A']:
                if cell.value == '':
                    cell.value = 'unchanged'           
                                    
        
        if only_in_oldsheet != None:
            for item in only_in_oldsheet:
                sheet_name=item  
                print('sheet only in old file: %s' % (sheet_name))
                old_ws = old_wb[sheet_name]
                new_ws = new_wb.create_sheet(title = sheet_name)
                for row in old_ws.iter_rows():
                    for cell in row:
                        new_ws.cell(row = cell.row, column = cell.column, value = cell.value)
                        new_ws.cell(cell.row, cell.column).fill = copy.copy(cell.fill)
                        new_ws.cell(cell.row, cell.column).font = copy.copy(cell.font)
                        new_ws.cell(cell.row, cell.column).border = copy.copy(cell.border)
                        
                create_Row_status(new_ws)  
                
                for cell in new_ws['A']:
                    if cell.row != 1:
                        cell.value = 'deleted'
                                 
        if only_in_newsheet != None:
            for item in only_in_newsheet:
                sheet_name = item
                print('sheet only in new file: %s' % (sheet_name))
                new_ws = new_wb[sheet_name]
                
                create_Row_status(new_ws) 
                
                for cell in new_ws['A']:
                    if cell.row != 1:
                        cell.value = 'newly added'
                
        new_wb.save(self.compfile_path)